import streamlit as st
import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
import time
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
import re

# Page configuration
st.set_page_config(
    page_title="EMA Alignment Scanner",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Define stock markets data
us_indices = {
    'S&P 500': '^GSPC',
    'Dow Jones': '^DJI',
    'NASDAQ': '^IXIC'
}

india_indices = {
    'NIFTY 50': '^NSEI',
    'SENSEX': '^BSESN',
    'NIFTY BANK': '^NSEBANK'
}

# Function to sanitize symbols
def sanitize_symbol(symbol):
    """Sanitize stock symbols to prevent injection attacks"""
    if not isinstance(symbol, str):
        return ""
    
    # Allow only alphanumeric characters, dots, hyphens, and ^ for indices
    sanitized = re.sub(r'[^A-Za-z0-9.\-^]', '', str(symbol).strip())
    
    # Limit length to prevent abuse
    sanitized = sanitized[:20]
    
    return sanitized

# Function to sanitize company names
def sanitize_name(name):
    """Sanitize company names to prevent injection attacks"""
    if not isinstance(name, str):
        return ""
    
    # Allow alphanumeric, spaces, common punctuation
    sanitized = re.sub(r'[^A-Za-z0-9\s\.,&\-\(\)]', '', str(name).strip())
    
    # Limit length to prevent abuse
    sanitized = sanitized[:200]
    
    return sanitized

# Function to load stock lists
@st.cache_data(ttl=86400)
def load_stock_lists():
    # Load US Stocks from Excel
    try:
        us_stocks = pd.read_excel('data/us_stocks.xlsx')
        if not all(col in us_stocks.columns for col in ['Symbol', 'Company Name']):
            # Try alternative column names
            column_mapping = {}
            for col in us_stocks.columns:
                if col.lower() in ['symbol', 'ticker', 'stock']:
                    column_mapping[col] = 'Symbol'
                elif col.lower() in ['name', 'company', 'company name', 'stock name']:
                    column_mapping[col] = 'Company Name'
            
            if column_mapping:
                us_stocks = us_stocks.rename(columns=column_mapping)
    except Exception as e:
        st.warning(f"Failed to load US stocks Excel: {e}. Using default list.")
        us_stocks = pd.DataFrame({
            'Symbol': ['AAPL', 'MSFT', 'AMZN', 'GOOGL', 'META', 'TSLA', 'NVDA', 'JPM', 'V', 'WMT'],
            'Company Name': ['Apple', 'Microsoft', 'Amazon', 'Alphabet', 'Meta Platforms', 'Tesla', 'NVIDIA', 'JPMorgan Chase', 'Visa', 'Walmart']
        })
    
    # Load Indian Stocks from Excel
    try:
        india_stocks = pd.read_excel('data/india_stocks.xlsx')
        if not all(col in india_stocks.columns for col in ['Symbol', 'Company Name']):
            # Try alternative column names
            column_mapping = {}
            for col in india_stocks.columns:
                if col.lower() in ['symbol', 'ticker', 'stock']:
                    column_mapping[col] = 'Symbol'
                elif col.lower() in ['name', 'company', 'company name', 'stock name']:
                    column_mapping[col] = 'Company Name'
            
            if column_mapping:
                india_stocks = india_stocks.rename(columns=column_mapping)
        
        # Ensure Indian stock symbols have .NS suffix for API calls
        india_stocks['Symbol'] = india_stocks['Symbol'].apply(
            lambda x: sanitize_symbol(x) if str(x).endswith('.NS') else f"{sanitize_symbol(x)}.NS"
        )
    except Exception as e:
        st.warning(f"Failed to load India stocks Excel: {e}. Using default list.")
        india_stocks = pd.DataFrame({
            'Symbol': ['RELIANCE.NS', 'TCS.NS', 'HDFCBANK.NS', 'INFY.NS', 'ICICIBANK.NS', 
                     'HINDUNILVR.NS', 'ITC.NS', 'SBIN.NS', 'BAJFINANCE.NS', 'BHARTIARTL.NS'],
            'Company Name': ['Reliance Industries', 'Tata Consultancy Services', 'HDFC Bank', 'Infosys', 
                    'ICICI Bank', 'Hindustan Unilever', 'ITC', 'State Bank of India', 
                    'Bajaj Finance', 'Bharti Airtel']
        })
    
    # Sanitize all symbols and names
    us_stocks['Symbol'] = us_stocks['Symbol'].apply(sanitize_symbol)
    us_stocks['Company Name'] = us_stocks['Company Name'].apply(sanitize_name)
    india_stocks['Symbol'] = india_stocks['Symbol'].apply(sanitize_symbol)
    india_stocks['Company Name'] = india_stocks['Company Name'].apply(sanitize_name)
    
    # Remove empty entries
    us_stocks = us_stocks[(us_stocks['Symbol'].str.len() > 0) & (us_stocks['Company Name'].str.len() > 0)]
    india_stocks = india_stocks[(india_stocks['Symbol'].str.len() > 0) & (india_stocks['Company Name'].str.len() > 0)]
    
    return us_stocks, india_stocks

# Function to process uploaded stock list
def process_uploaded_stock_list(uploaded_file, market):
    try:
        # Read Excel file only
        if not uploaded_file.name.endswith('.xlsx'):
            st.error("Only Excel (.xlsx) files are supported")
            return None
            
        stocks_df = pd.read_excel(uploaded_file)
        
        # Standardize column names (case-insensitive)
        column_mapping = {}
        for col in stocks_df.columns:
            if col.lower() in ['symbol', 'ticker', 'stock']:
                column_mapping[col] = 'Symbol'
            elif col.lower() in ['name', 'company', 'company name', 'stock name']:
                column_mapping[col] = 'Company Name'
        
        # Rename columns if needed
        if column_mapping:
            stocks_df = stocks_df.rename(columns=column_mapping)
        
        # Check if we have the required columns
        if 'Symbol' not in stocks_df.columns:
            raise ValueError("File must contain a 'Symbol' column")
        
        # If no name column exists, create one with symbol values
        if 'Company Name' not in stocks_df.columns:
            stocks_df['Company Name'] = stocks_df['Symbol']
        
        # Sanitize all data
        stocks_df['Symbol'] = stocks_df['Symbol'].apply(sanitize_symbol)
        stocks_df['Company Name'] = stocks_df['Company Name'].apply(sanitize_name)
        
        # Remove empty entries
        stocks_df = stocks_df[(stocks_df['Symbol'].str.len() > 0) & (stocks_df['Company Name'].str.len() > 0)]
        
        # Ensure proper formatting for Indian stocks
        if market == "India":
            stocks_df['Symbol'] = stocks_df['Symbol'].apply(
                lambda x: x if str(x).endswith('.NS') else f"{x}.NS"
            )
        
        # Limit to 9999 stocks
        if len(stocks_df) > 9999:
            stocks_df = stocks_df.iloc[:9999]
            st.warning(f"Stock list limited to 9999 stocks")
        
        return stocks_df
        
    except Exception as e:
        st.error(f"Error processing uploaded file: {e}")
        return None

# Function to get stock data and calculate EMAs
@st.cache_data(ttl=3600)
def get_stock_data(symbol, timeframe):
    try:
        # Sanitize symbol before API call
        symbol = sanitize_symbol(symbol)
        if not symbol:
            return None
            
        stock = yf.Ticker(symbol)
        
        # Set period based on timeframe
        if timeframe == "1d":
            period = "500d"
        elif timeframe == "15m":
            period = "30d"
        elif timeframe == "1wk":
            period = "7y"
        else:  # 1h
            period = "90d"
            
        df = stock.history(period=period, interval=timeframe)
        
        if df.empty or len(df) < 200:  # Ensure we have enough data for EMAs
            return None
        
        # Calculate EMAs precisely
        df['EMA20'] = df['Close'].ewm(span=20, adjust=False).mean()
        df['EMA50'] = df['Close'].ewm(span=50, adjust=False).mean()
        df['EMA100'] = df['Close'].ewm(span=100, adjust=False).mean()
        df['EMA200'] = df['Close'].ewm(span=200, adjust=False).mean()
        
        return df
    except Exception as e:
        return None

# Function to check EMA alignment
def check_ema_alignment(df):
    if df is None or df.empty:
        return None, None
    
    # Get the latest values
    latest = df.iloc[-1]
    close_price = latest['Close']
    ema20 = latest['EMA20']
    ema50 = latest['EMA50']
    ema100 = latest['EMA100']
    ema200 = latest['EMA200']
    
    # Check bullish alignment: Close > EMA20 > EMA50 > EMA100 > EMA200
    is_bullish = (close_price > ema20 > ema50 > ema100 > ema200)
    
    # Check bearish alignment: Close < EMA20 < EMA50 < EMA100 < EMA200
    is_bearish = (close_price < ema20 < ema50 < ema100 < ema200)
    
    if is_bullish:
        return "Bullish", "🟢"
    elif is_bearish:
        return "Bearish", "🔴"
    else:
        return None, None

# Function to scan all stocks for EMA alignment
def scan_ema_alignment(stock_list, timeframe, market):
    results = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_stocks = len(stock_list)
    processed_count = 0
    
    for i, (symbol, name) in enumerate(zip(stock_list['Symbol'], stock_list['Company Name'])):
        status_text.text(f"Scanning {market} stocks: {i+1}/{total_stocks} - {name} ({symbol})")
        progress_bar.progress((i + 1) / total_stocks)
        
        df = get_stock_data(symbol, timeframe)
        
        if df is None or df.empty:
            continue
            
        processed_count += 1
        
        trend, status_emoji = check_ema_alignment(df)
        
        if trend:  # Only add if bullish or bearish alignment found
            # Remove .NS suffix and ^ symbol for display
            display_symbol = symbol.replace('.NS', '') if symbol.endswith('.NS') else symbol
            display_symbol = display_symbol.replace('^', '') if display_symbol.startswith('^') else display_symbol
            
            results.append({
                'Symbol': display_symbol,
                'Company Name': name,
                'Trend': trend,
                'Status': status_emoji,
                'Original_Symbol': symbol  # Keep original for any further processing
            })
    
    progress_bar.empty()
    status_text.empty()
    
    # Show summary of scan results
    if processed_count < total_stocks:
        st.info(f"Note: Data for {total_stocks - processed_count} stocks could not be retrieved or processed.")
    
    return pd.DataFrame(results) if results else pd.DataFrame()

# Function to create formatted Excel file - FIXED VERSION
def create_formatted_excel(df, filename):
    if df.empty:
        return None
    
    # Create a copy of dataframe for export (without Original_Symbol)
    export_df = df[['Symbol', 'Company Name', 'Trend', 'Status']].copy()
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    # Use xlsxwriter engine for better Excel compatibility
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            export_df.to_excel(writer, sheet_name='EMA Alignment Results', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['EMA Alignment Results']
            
            # Define colors and fills
            green_font = Font(color="00008000", bold=True)  # Green
            red_font = Font(color="00FF0000", bold=True)    # Red
            green_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # Light green background
            red_fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")    # Light red background
            
            # Format the data rows
            for row in range(2, len(export_df) + 2):  # Start from row 2, skip header
                trend_value = worksheet[f'C{row}'].value
                if trend_value == 'Bullish':
                    # Color the entire row green for bullish stocks
                    for col in ['A', 'B', 'C', 'D']:
                        cell = worksheet[f'{col}{row}']
                        cell.font = green_font
                        cell.fill = green_fill
                        # Fix the status emoji display issue
                        if col == 'D':
                            cell.value = "Bullish"
                elif trend_value == 'Bearish':
                    # Color the entire row red for bearish stocks
                    for col in ['A', 'B', 'C', 'D']:
                        cell = worksheet[f'{col}{row}']
                        cell.font = red_font
                        cell.fill = red_fill
                        # Fix the status emoji display issue
                        if col == 'D':
                            cell.value = "Bearish"
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    except Exception as e:
        st.error(f"Error creating Excel file: {e}")
        return None
    
    output.seek(0)
    return output

# Main application
def main():
    st.title("EMA Alignment Scanner")
    
    # Display current market status at the top
    st.subheader("Market Status")
    col1, col2, col3 = st.columns(3)
    
    # Initialize session state for managing stock lists
    if 'using_custom_list' not in st.session_state:
        st.session_state.using_custom_list = False
    
    # Sidebar
    st.sidebar.header("Scanner Settings")
    
    # Load default stock lists
    us_stocks, india_stocks = load_stock_lists()
    
    # Custom stock list upload
    st.sidebar.subheader("Stock List")
    uploaded_file = st.sidebar.file_uploader(
        "Upload Custom (Symbol, Company Name)",
        type=["xlsx"],
        help="Excel file with 'Symbol' and 'Company Name' columns (Max 50MB, 9999 stocks)"
    )
    
    # Process uploaded file if available
    custom_stocks = None
    if uploaded_file is not None:
        if uploaded_file.size > 50 * 1024 * 1024:  # 50MB limit
            st.sidebar.error("File size exceeds 50MB limit")
            st.session_state.using_custom_list = False
        else:
            market_for_processing = st.session_state.get('market', "India")
            custom_stocks = process_uploaded_stock_list(uploaded_file, market_for_processing)
            
            if custom_stocks is not None:
                st.session_state.using_custom_list = True
                st.session_state.custom_stocks = custom_stocks
                st.sidebar.success(f"Loaded {len(custom_stocks)} stocks from your file")
            else:
                st.session_state.using_custom_list = False
    else:
        st.session_state.using_custom_list = False
    
    # Market selection - Default to India
    if st.session_state.using_custom_list:
        market = st.sidebar.selectbox(
            "Select Market (Disabled - Using Custom List)",
            ["India", "US"],
            disabled=True,
            index=0 if st.session_state.get('market', "India") == "India" else 1
        )
        market = st.session_state.get('market', "India")
    else:
        market = st.sidebar.selectbox("Select Market", ["India", "US"])
        st.session_state.market = market
    
    # Timeframe selection - Updated with new timeframes, Default to Daily
    timeframe_options = {
        "Daily": "1d",
        "Hourly": "1h",
        "15 Minutes": "15m",
        "Weekly": "1wk"
    }
    timeframe_display = st.sidebar.selectbox("Select Timeframe", list(timeframe_options.keys()), index=0)
    timeframe = timeframe_options[timeframe_display]
    
    # Scan button
    scan_button = st.sidebar.button("Start EMA Alignment Scan", use_container_width=True)
    
    # Display current market status data
    indices = india_indices if market == "India" else us_indices
    
    index_cols = [col1, col2, col3]
    for i, (index_name, index_symbol) in enumerate(indices.items()):
        try:
            # Sanitize index symbol
            sanitized_index_symbol = sanitize_symbol(index_symbol)
            index_data = yf.Ticker(sanitized_index_symbol).history(period="1d")
            if not index_data.empty:
                current = index_data['Close'].iloc[-1]
                previous = index_data['Open'].iloc[-1]
                change = current - previous
                change_percent = (change / previous) * 100
                
                color = "green" if change >= 0 else "red"
                change_icon = "▲" if change >= 0 else "▼"
                
                index_cols[i].markdown(
                    f"**{index_name}**: {current:.2f} "
                    f"<span style='color:{color}'>{change_icon} {abs(change):.2f} ({abs(change_percent):.2f}%)</span>", 
                    unsafe_allow_html=True
                )
        except:
            index_cols[i].text(f"{index_name}: Data unavailable")
    
    if scan_button:
        # Use custom stock list if uploaded, otherwise use default
        if st.session_state.using_custom_list:
            stocks_to_scan = st.session_state.custom_stocks
        else:
            stocks_to_scan = india_stocks if market == "India" else us_stocks
        
        with st.spinner(f"Scanning {market} stocks for EMA alignment on {timeframe_display} timeframe..."):
            results_df = scan_ema_alignment(stocks_to_scan, timeframe, market)
        
        # Store results in session state
        st.session_state.results_df = results_df
        st.session_state.last_scan_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        st.session_state.market = market
        st.session_state.timeframe = timeframe_display
    
    # Display explanation
    with st.expander("How This EMA Alignment Scanner Works"):
        st.markdown(f"""
        **Disclaimer**: This project is intended for educational and informational purposes only. You are solely responsible for any profits or losses you may incur.
        
        ## EMA Alignment Scanner Logic
        
        This scanner identifies stocks with perfect EMA alignment:
        
        ### 🟢 Bullish Stocks
        - Latest Close Price > EMA 20 > EMA 50 > EMA 100 > EMA 200
        - Perfect bullish alignment indicates strong upward momentum
        
        ### 🔴 Bearish Stocks  
        - Latest Close Price < EMA 20 < EMA 50 < EMA 100 < EMA 200
        - Perfect bearish alignment indicates strong downward momentum
        
        ### Timeframes Available
        - **Daily**: Uses 500 days of data for mid-term analysis
        - **Hourly**: Uses 90 days of data for swing analysis
        - **15 Minutes**: Uses 30 days of data for short-term analysis
        - **Weekly**: Uses 7 years of data for long-term analysis
        
        ### Important Notes
        - All EMAs are calculated precisely using exponential weighting
        - Only stocks with perfect alignment are shown
        - Indian stock symbols display without .NS suffix in results
        - Export files are formatted with color coding (Green for Bullish, Red for Bearish)
        - All data is sanitized for security
        
        ### Using Custom Stock Lists
        - Upload Excel files with 'Symbol' and 'Company Name' columns
        - Maximum 9999 stocks per list and 50MB file size
        - For Indian stocks, .NS suffix is automatically handled
        """)
    
    # Display results
    if 'results_df' in st.session_state and not st.session_state.results_df.empty:
        st.subheader("EMA Alignment Results")
        
        # Show last scan info
        scan_info = f"Last scan: {st.session_state.last_scan_time} | Market: {st.session_state.market} | Timeframe: {st.session_state.timeframe}"
        st.info(scan_info)
        
        # Separate bullish and bearish stocks
        bullish_stocks = st.session_state.results_df[st.session_state.results_df['Trend'] == 'Bullish']
        bearish_stocks = st.session_state.results_df[st.session_state.results_df['Trend'] == 'Bearish']
        
        # Create tabs for bullish and bearish
        tab1, tab2 = st.tabs([f"Bullish Stocks 🟢 ({len(bullish_stocks)})", f"Bearish Stocks 🔴 ({len(bearish_stocks)})"])
        
        with tab1:
            if not bullish_stocks.empty:
                st.subheader("Perfect Bullish EMA Alignment")
                display_df = bullish_stocks[['Symbol', 'Company Name', 'Trend', 'Status']].copy()
                st.dataframe(display_df, use_container_width=True)
                
                # Download button for bullish stocks - FIXED
                excel_file = create_formatted_excel(bullish_stocks, f"bullish_stocks_{st.session_state.market}_{st.session_state.timeframe}")
                if excel_file:
                    st.download_button(
                        label="📥 Download Bullish Stocks (Excel)",
                        data=excel_file,
                        file_name=f"bullish_stocks_{st.session_state.market}_{st.session_state.timeframe}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.info("No stocks found with perfect bullish EMA alignment.")
        
        with tab2:
            if not bearish_stocks.empty:
                st.subheader("Perfect Bearish EMA Alignment")
                display_df = bearish_stocks[['Symbol', 'Company Name', 'Trend', 'Status']].copy()
                st.dataframe(display_df, use_container_width=True)
                
                # Download button for bearish stocks - FIXED
                excel_file = create_formatted_excel(bearish_stocks, f"bearish_stocks_{st.session_state.market}_{st.session_state.timeframe}")
                if excel_file:
                    st.download_button(
                        label="📥 Download Bearish Stocks (Excel)",
                        data=excel_file,
                        file_name=f"bearish_stocks_{st.session_state.market}_{st.session_state.timeframe}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.info("No stocks found with perfect bearish EMA alignment.")
        
        # Download all results button - FIXED
        if not st.session_state.results_df.empty:
            st.subheader("Download All Results")
            excel_file = create_formatted_excel(st.session_state.results_df, f"ema_alignment_results_{st.session_state.market}_{st.session_state.timeframe}")
            if excel_file:
                st.download_button(
                    label="📥 Download All Results (Excel)",
                    data=excel_file,
                    file_name=f"ema_alignment_results_{st.session_state.market}_{st.session_state.timeframe}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    elif 'last_scan_time' in st.session_state:
        st.info("No stocks found with perfect EMA alignment. Try scanning with different parameters.")
    else:
        st.info("Click 'Start EMA Alignment Scan' to begin scanning for stocks with perfect EMA alignment.")

# Run the application
if __name__ == "__main__":
    main()